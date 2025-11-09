"""
Custom Excel Reader Library for Robot Framework
Author: ClearTrip Automation Team
Description: A custom library to read Excel files and return data for Robot Framework tests
"""

import openpyxl
import pandas as pd
from robot.api.deco import keyword
from robot.api import logger
from typing import List, Dict, Any
import os


class ExcelReader:
    """
    Custom Excel Reader Library for Robot Framework
    
    This library provides various methods to read Excel files and return data
    in formats suitable for Robot Framework data-driven testing.
    """

    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    ROBOT_LIBRARY_VERSION = '1.0.0'

    def __init__(self):
        """Initialize the ExcelReader library."""
        self.current_workbook = None
        self.current_worksheet = None

    @keyword("Read Excel Data")
    def read_excel_data(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Read Excel file and return data as list of dictionaries.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of dictionaries where each dictionary represents a row
            
        Example:
            | ${data} | Read Excel Data | path/to/file.xlsx |
        """
        logger.info(f"Reading Excel file: {file_path}")
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        try:
            # Using openpyxl to read Excel file
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            # Get data rows
            data_rows = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                data_rows.append(row_dict)
            
            workbook.close()
            logger.info(f"Successfully read {len(data_rows)} rows from Excel file")
            return data_rows
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise

    @keyword("Read Excel Data With Pandas")
    def read_excel_data_with_pandas(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Read Excel file using pandas library.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of dictionaries where each dictionary represents a row
            
        Example:
            | ${data} | Read Excel Data With Pandas | path/to/file.xlsx |
        """
        logger.info(f"Reading Excel file with pandas: {file_path}")
        
        try:
            # Using pandas to read Excel file
            df = pd.read_excel(file_path)
            
            # Convert NaN values to empty strings
            df = df.fillna('')
            
            # Convert to list of dictionaries
            data_list = df.to_dict('records')
            
            logger.info(f"Successfully read {len(data_list)} rows with pandas")
            return data_list
            
        except Exception as e:
            logger.error(f"Error reading Excel file with pandas: {str(e)}")
            raise

    @keyword("Read Excel Sheet By Name")
    def read_excel_sheet_by_name(self, file_path: str, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Read specific sheet from Excel file.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read
            
        Returns:
            List of dictionaries where each dictionary represents a row
            
        Example:
            | ${data} | Read Excel Sheet By Name | path/to/file.xlsx | Sheet1 |
        """
        logger.info(f"Reading sheet '{sheet_name}' from Excel file: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            if sheet_name not in workbook.sheetnames:
                available_sheets = ", ".join(workbook.sheetnames)
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
            
            worksheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            # Get data rows
            data_rows = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                data_rows.append(row_dict)
            
            workbook.close()
            logger.info(f"Successfully read {len(data_rows)} rows from sheet '{sheet_name}'")
            return data_rows
            
        except Exception as e:
            logger.error(f"Error reading Excel sheet: {str(e)}")
            raise

    @keyword("Get Excel Sheet Names")
    def get_excel_sheet_names(self, file_path: str) -> List[str]:
        """
        Get list of sheet names from Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of sheet names
            
        Example:
            | ${sheets} | Get Excel Sheet Names | path/to/file.xlsx |
        """
        logger.info(f"Getting sheet names from Excel file: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            logger.info(f"Found sheets: {sheet_names}")
            return sheet_names
            
        except Exception as e:
            logger.error(f"Error getting sheet names: {str(e)}")
            raise

    @keyword("Read Excel Cell Value")
    def read_excel_cell_value(self, file_path: str, row: int, column: int, sheet_name: str = None):
        """
        Read specific cell value from Excel file.
        
        Args:
            file_path: Path to the Excel file
            row: Row number (1-based)
            column: Column number (1-based)
            sheet_name: Sheet name (optional, uses active sheet if not provided)
            
        Returns:
            Cell value
            
        Example:
            | ${value} | Read Excel Cell Value | path/to/file.xlsx | 2 | 3 |
            | ${value} | Read Excel Cell Value | path/to/file.xlsx | 2 | 3 | Sheet1 |
        """
        logger.info(f"Reading cell value from row {row}, column {column}")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            if sheet_name:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            cell_value = worksheet.cell(row=row, column=column).value
            workbook.close()
            
            logger.info(f"Cell value: {cell_value}")
            return cell_value
            
        except Exception as e:
            logger.error(f"Error reading cell value: {str(e)}")
            raise

    @keyword("Validate Excel File Structure")
    def validate_excel_file_structure(self, file_path: str, expected_columns: List[str]) -> bool:
        """
        Validate that Excel file has expected column structure.
        
        Args:
            file_path: Path to the Excel file
            expected_columns: List of expected column names
            
        Returns:
            True if structure matches, False otherwise
            
        Example:
            | ${is_valid} | Validate Excel File Structure | path/to/file.xlsx | test_name,from_city,to_city |
        """
        logger.info(f"Validating Excel file structure: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Get actual headers
            actual_headers = []
            for cell in worksheet[1]:
                actual_headers.append(cell.value)
            
            workbook.close()
            
            # Compare headers
            headers_match = set(actual_headers) == set(expected_columns)
            
            if headers_match:
                logger.info("Excel file structure validation passed")
            else:
                logger.warn(f"Excel file structure mismatch. Expected: {expected_columns}, Actual: {actual_headers}")
            
            return headers_match
            
        except Exception as e:
            logger.error(f"Error validating Excel file structure: {str(e)}")
            return False